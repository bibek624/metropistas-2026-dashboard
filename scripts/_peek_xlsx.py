# -*- coding: utf-8 -*-
import os
try:
    import openpyxl
except ImportError:
    print("NO_OPENPYXL")
    raise SystemExit

paths = [
    r"D:\Projects\Metropistas_2026\Raw Data\020 MFV\MFV_Length_Summary.xlsx",
    r"D:\Projects\Metropistas_2026\Raw Data\Puerto Rico Tracking Year 3-2026-Final_BP.xlsx",
]
for p in paths:
    if not os.path.exists(p):
        print("MISSING", p)
        continue
    print("\n############", os.path.basename(p), "############")
    wb = openpyxl.load_workbook(p, data_only=True)
    for ws in wb.worksheets:
        print("=== sheet: %s (%dx%d) ===" % (ws.title, ws.max_row, ws.max_column))
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 35:
                print("   ...more rows...")
                break
            if all(c is None for c in row):
                continue
            cells = [round(c, 2) if isinstance(c, float) else c for c in row]
            print("  ", cells)
